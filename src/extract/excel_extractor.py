"""
Extraction des données de production depuis les exports Excel.

Responsabilité UNIQUE de ce module : lire un fichier Excel et retourner un
DataFrame brut de la feuille "Détails Cadences", enrichi de métadonnées de
traçabilité. Il ne fait NI contrôle qualité, NI transformation métier,
NI calcul — ces responsabilités appartiennent respectivement aux modules
quality/, transform/ et analytics/ (Phases 3, 5+).

Ce découpage strict est ce qui permettra au pipeline de rester maintenable
quand il grossira : chaque module peut être testé, modifié et remplacé
indépendamment des autres.
"""

from __future__ import annotations

import logging
from dataclasses import dataclass, field
from datetime import datetime
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from src.extract.exceptions import (
    CorruptedFileError,
    EmptySheetError,
    FileNotFoundExtractionError,
    SheetNotFoundError,
)
from src.utils.config import AppConfig


@dataclass
class ExtractionResult:
    """Enveloppe le DataFrame extrait + les métadonnées de l'extraction."""

    dataframe: pd.DataFrame
    source_file: Path
    sheet_name: str
    n_rows: int
    n_columns: int
    extracted_at: datetime = field(default_factory=datetime.now)

    def summary(self) -> str:
        return (
            f"Extraction de '{self.source_file.name}' → feuille "
            f"'{self.sheet_name}' : {self.n_rows} lignes × "
            f"{self.n_columns} colonnes."
        )


class ExcelExtractor:
    """
    Extrait la feuille "Détails Cadences" d'un export Excel de production.

    Usage :
        extractor = ExcelExtractor(config)
        result = extractor.extract("data/raw/export_2026_06.xlsx")
        df = result.dataframe
    """

    def __init__(self, config: AppConfig, logger: logging.Logger | None = None):
        self.config = config
        self.logger = logger or logging.getLogger("cadence_pipeline")

    # ------------------------------------------------------------------
    # API publique
    # ------------------------------------------------------------------

    def list_sheets(self, file_path: str | Path) -> list[str]:
        """Liste les feuilles disponibles dans le classeur, sans le charger entièrement."""
        path = self._validate_file_exists(file_path)
        try:
            workbook = load_workbook(path, read_only=True, data_only=True)
            sheets = workbook.sheetnames
            workbook.close()
            return sheets
        except Exception as exc:  # noqa: BLE001 - on re-raise typé ci-dessous
            raise CorruptedFileError(
                f"Impossible de lire le classeur '{path.name}' : {exc}"
            ) from exc

    def extract(self, file_path: str | Path) -> ExtractionResult:
        """
        Extrait la feuille "Détails Cadences" du fichier donné.

        Lève :
            FileNotFoundExtractionError si le fichier n'existe pas.
            SheetNotFoundError si la feuille requise est absente.
            CorruptedFileError si le fichier ne peut pas être ouvert.
            EmptySheetError si la feuille est vide.
        """
        path = self._validate_file_exists(file_path)
        target_sheet = self.config.excel.required_sheet

        available_sheets = self.list_sheets(path)
        if target_sheet not in available_sheets:
            raise SheetNotFoundError(
                f"Feuille '{target_sheet}' introuvable dans '{path.name}'. "
                f"Feuilles disponibles : {available_sheets}"
            )

        try:
            df = pd.read_excel(
                path,
                sheet_name=target_sheet,
                header=self.config.excel.header_row,
                engine="openpyxl",
            )
        except Exception as exc:  # noqa: BLE001
            raise CorruptedFileError(
                f"Échec de lecture de la feuille '{target_sheet}' dans "
                f"'{path.name}' : {exc}"
            ) from exc

        if df.empty:
            raise EmptySheetError(
                f"La feuille '{target_sheet}' de '{path.name}' ne contient aucune donnée."
            )

        # Nettoyage minimal non métier : suppression des colonnes 'Unnamed: X'
        # générées par openpyxl quand l'export Excel contient des cellules
        # fusionnées ou des colonnes vides en en-tête.
        df = df.loc[:, ~df.columns.astype(str).str.startswith("Unnamed:")]

        if self.config.extraction.add_technical_columns:
            df = self._add_technical_columns(df, path)

        result = ExtractionResult(
            dataframe=df,
            source_file=path,
            sheet_name=target_sheet,
            n_rows=len(df),
            n_columns=len(df.columns),
        )
        self.logger.info(result.summary())
        return result

    # ------------------------------------------------------------------
    # Méthodes privées
    # ------------------------------------------------------------------

    def _validate_file_exists(self, file_path: str | Path) -> Path:
        path = Path(file_path)
        if not path.exists():
            raise FileNotFoundExtractionError(f"Fichier introuvable : {path}")
        if path.suffix.lower() not in (".xlsx", ".xlsm"):
            raise CorruptedFileError(
                f"Extension non supportée : '{path.suffix}'. Attendu : .xlsx ou .xlsm"
            )
        return path

    def _add_technical_columns(self, df: pd.DataFrame, source_path: Path) -> pd.DataFrame:
        """
        Ajoute des colonnes techniques de traçabilité, indispensables pour
        la Phase 4 (historisation) : savoir de quel fichier et quand chaque
        ligne provient, et pouvoir déduplicater les imports.
        """
        df = df.copy()
        df["_source_file"] = source_path.name
        df["_extracted_at"] = datetime.now().isoformat()
        return df
